#!/usr/bin/env python3
# 
# XMerge.py - 2022,23 by Gregory A. Sanders (dr.gerg@drgerg.com)
# Merge multiple source files using a common header into one output file.
#

import tkinter as tk
import openpyxl as xl
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from tkinter.font import Font
import time
import sys
from configparser import ConfigParser
from os import path
import os
import warnings
import unicodedata
import csv
import pathlib
import shutil
from tkhtmlview import HTMLScrolledText, RenderHTML, HTMLLabel
import argparse

version = "v1.4.2"
# Changes in v1.4.2:
#       1. Handle multiple input files with the same filename.
#       2. Add the ability to save the output file as .csv.
#       3. Add CLI args -x and -c to control output format.
#
# NOTES: XMerge.py uses (2) .ini files via the configparser module.
#
#       "xmerge.ini" is the system-wide config file.  It is kept in the 
#       installation folder.
#       The xmerge.ini file tells XMerge.py where to start, specifically, 
#       what was the last output folder XMerge.py worked in.
#
#       "LastXMerge.ini" is a per-output-folder config file.  It is located
#       in each output folder, and contains the paths and filenames
#       used to create that output (export) file.
#
#       We differentiate between the two .ini files by making separate
#       instances of ConfigParser(): 'confparse' for LastXMerge.ini
#       and 'sysconfparse' for the system-wide xmerge.ini.
#
xmargparse = argparse.ArgumentParser()
xmargparse.add_argument(
    "-u",
    "--unattended",
    help="Run XMerge without interaction. Provide full path (including backslash) to LastXMerge.ini file in quotes.",
    action="store",
)
xmargparse.add_argument(
    "-c",
    "--csv",
    help="Create a .csv export in addition to the .xlsx export.",
    action="store_true",
)
xmargparse.add_argument(
    "-x",
    "--xlsx",
    help="DO NOT create a .xlsx export.",
    action="store_true",
)
clarg = xmargparse.parse_args()
if clarg.unattended:
    if clarg.unattended[:-1] != "\\":
        clarg.unattended = clarg.unattended + "\\"
    u_outfolder = os.path.dirname(clarg.unattended)


confparse = ConfigParser()
sysconfparse = ConfigParser()
path_to_dat = path.abspath(path.dirname(__file__))
XMergeIni = path_to_dat + "\\xmerge.ini"
sysconfparse.read(XMergeIni)
outFolderChk = sysconfparse.get('folders', 'output_folder')
#
# SET WORKING ROOT AND EXPORTS FOLDER LOCATIONS
#
def setup():
    path_to_dat = path.abspath(path.dirname(__file__))
    XMergeIni = path_to_dat + "\\xmerge.ini"
    userHome = os.path.expanduser("~")
    datapath = ""
    # Check the xmerge.ini file for previously visited folders.
    # Get previous Data folder.
    # Then go to that output folder (if it exists) and get LastXMerge.ini to process.
    sysconfparse.read(XMergeIni)
    # IF xmerge.ini has inaccurate data, edit it.
    # 'sysconfparse' is for xmerge.ini.
    if clarg.unattended:
        outFolderChk = u_outfolder
    else:
        outFolderChk = sysconfparse.get('folders', 'output_folder')
    ofExist = os.path.isdir(outFolderChk)
    didEdit = 0
    if ofExist == False:            # Configured output folder isn't there.
        outFilesChkBox.select()     # Check the 'Change Output Folder' checkbox.
        outxm = 1
        sysconfparse.set('folders', 'output_folder', "")
        didEdit = 1
    instchk = sysconfparse.get('folders', 'installation')
    homepath = sysconfparse.get('folders', 'homepath')
    verchk = sysconfparse.get('program','version')
    if instchk != path_to_dat:
        sysconfparse.set('folders', 'installation', path_to_dat)
        didEdit = 1
    if homepath != userHome:
        sysconfparse.set('folders', 'homepath', userHome)
        didEdit = 1
    if verchk != version:
        sysconfparse.set('program','version',version)
        didEdit = 1
    if outFolderChk == "":
        outFilesChkBox.select()
        outFolderChk = newOutFolder()
    else:
        outFilesChkBox.deselect()
    if didEdit == 1:
        with open(XMergeIni, 'w') as nnIni:
            sysconfparse.write(nnIni)
            didEdit = 0
    #
    # SHIFT TO LastXMerge.ini to get the rest of this.
    # If it doesn't exist, make one.
    # 'confparse' is for LastXMerge.ini.
    #
    if clarg.unattended:
        LastMergeIni = clarg.unattended + "\\LastXMerge.ini"
        outFolderChk = u_outfolder
    else:
        outFolderChk = sysconfparse.get('folders', 'output_folder') # may have changed.
        LastMergeIni = outFolderChk + "\\LastXMerge.ini"
    iniFilesList = []                               # Empty list for filenames.
    chkINI = os.path.isfile(LastMergeIni)           # Look for the .ini file.
    if chkINI == True:
        confparse.read(LastMergeIni)
         ## Update the LastXMerge.ini file if it has archaic options.
        if confparse.has_option('userselections','columnselection') == False:
            confparse.set('userselections','columnselection','')
        if confparse.has_option('userselections','status') == True:
            confparse.remove_option('userselections','status')
        confxportFn = confparse.get('export','exportfilename')   # get LastXMerge.ini value
        if confxportFn[-4:] != "xlsx":                           # normalize the filename with .xlsx.
            confxportFn = confxportFn + ".xlsx"
        xportFn.set(confxportFn)
        # upgrade .ini file to support multiple source folders
        if confparse.has_option('folders','data_folder'):
            dftemp = confparse.get('folders','data_folder')
            confparse.remove_option('folders','data_folder')
            confparse.set('folders','data_folder1',dftemp)
        datapath = confparse.get('folders', 'data_folder1')
        colNameChoice = confparse.get("userselections","columnselection")
        hdrRowChoice = confparse.get("userselections","headerrow")
        hdrrwFn.set(hdrRowChoice)
        appendfilename = confparse.get("userselections","appendfilename")
        aFnVar.set(int(appendfilename))
        if colNameChoice != "":
            colNameVar.set(colNameChoice)
        if confxportFn != xportFn:
            xportFn.set(confxportFn)
            with open(LastMergeIni, 'w') as LMIni:
                confparse.write(LMIni)
        iniFchk = {k:v for k,v in confparse['files'].items()}
        for k in iniFchk:
            iniFilesList.append(os.path.basename(iniFchk[k]))
        lastIniFound = " was "
    else:
        lastIniFound = " was NOT "
        newLastXMerge()
    text3.delete("1.0", tk.END)
    text3.insert(tk.INSERT, "Initialization complete. LastXMerge.ini" + lastIniFound + "found in Output Folder.")
    updateWinText()
    if clarg.unattended:
        main()
    #
    #


def updateWinText():
    # Everything we need to know has been gathered from the two .ini files.
    # Now we're ready to do something when the user hits a button.
    # Update the folder and path info in the GUI.
    #
    if clarg.unattended:
        outFolderChk = u_outfolder
    else:
        outFolderChk = sysconfparse.get('folders', 'output_folder')
    if clarg.csv:
        csvVar.set(1)
    if clarg.xlsx:
        xlsxVar.set(0)
    datapath = confparse.get('folders', 'data_folder1')
    datchk = {k:v for k,v in confparse['files'].items()}
    iniFilesList = []
    for k in datchk:
        iniFilesList.append(os.path.basename(datchk[k]))    # one for GUI
    text1.delete("1.0", tk.END)
    text1.insert(tk.INSERT, "Initial Source folder: " + datapath)
    text1.insert(tk.INSERT, "\nOutput folder: " + outFolderChk)
    if iniFilesList > []:
        text1.insert(tk.INSERT, "\nCurrently selected files: " + os.path.basename(str(iniFilesList).strip("[]")))
    window.update()
#
#  THE GO BUTTON BRINGS US HERE FIRST - THE BEGINNING - 
#
def main():
    ofcchk = getCtrlVals()[2]
    colNameChoice = getCtrlVals()[0]
    if ofcchk == 1:                 # Look for 'Clear Output Folder' checkmark.
        purgeExp()
    outxm = getCtrlVals()[1]        # Look for 'Change Output Folder' selection.
    if outxm == 1:
        outFolderChk = newOutFolder()
    else:
        if clarg.unattended:
            LastMergeIni = clarg.unattended + "\\LastXMerge.ini"
            outFolderChk = u_outfolder
        else:
            outFolderChk = sysconfparse.get('folders', 'output_folder') # may have changed.
            LastMergeIni = outFolderChk + "\\LastXMerge.ini"
    confparse.read(LastMergeIni)
    confxportFn = confparse.get('export','exportfilename')   # get LastXMerge.ini value
    CurrXportFilename = getCtrlVals()[3]
    if CurrXportFilename[-4:] != "xlsx":
        CurrXportFilename = CurrXportFilename + ".xlsx"
    # If the export filename is a new one, save it to the .ini file.
    if CurrXportFilename != confxportFn:
        confparse.set('export','exportfilename',CurrXportFilename)
        with open(LastMergeIni, 'w') as LMIni:
            confparse.write(LMIni)
        
    # Go get the source files now.
    firstFile,exportFolder,datchkFiles = browseFiles()  # Calls browseFiles(), then if Continue is Yes, txtFileCnvrt().
    # Copy the raw source files to the temp folder.
    exCOPYcontinue(exportFolder,datchkFiles)
    # Based on the source file types, make lists for conversions.
    xlsext, xlsxext, txtext, csvext, allEntr = initialize(exportFolder)
    # Go do the needed conversions from whatever to .xlsx.
    convertedList = []
    if xlsext != []:
        convertedList = xlsFileCnvrt(exportFolder,xlsext,convertedList)
    if txtext != []:
        convertedList = txtFileCnvrt(exportFolder,txtext,convertedList)
    if csvext != []:
        convertedList = csvFileCnvrt(exportFolder,csvext,convertedList)
    # .xlsx files get special handling because they don't require conversion.
    if xlsxext != []:           # We need to make a list, BUT
        convertedList = []      # we don't want to get the output file in the list.
        xpFnIni = confparse.get('export', 'exportfilename')
        for f in xlsxext:
            if f != xpFnIni:
                convertedList.append(exportFolder + "/temp/" + f)
    text3.delete('1.0',tk.END)
    text3.insert(tk.END, "Finished converting files in the Output Folder.")
    fileNam = convertedList[0] # The first .xlsx file will provide the header row
    if colNameChoice == 'all': # if 'All Columns' is selected.
        itemList = makeColsList(fileNam)
    elif colNameChoice == 'conf':
        itemList = colConfig()
    appendfilename = getCtrlVals()[5]
    outFolderChk = sysconfparse.get('folders', 'output_folder')
    if not clarg.unattended:
        confparse.set('userselections','appendfilename',str(appendfilename))
        LastMergeIni = outFolderChk + "\\LastXMerge.ini"
        with open(LastMergeIni, 'w') as LMIni:
            confparse.write(LMIni)
    if appendfilename == 1:
        itemList.append('SRCFILE')          # Append 'SRCFILE' to itemList so we can include a filename column.
    DESTfilename = makeOutput(itemList,exportFolder)
    # At this juncture, convertedList contains the .xlsx filenames of source files.
    # itemList contains the column names (defines the final output).
    # exportFolder is where we're putting all this output, and is where our source file copies live right now.
    # DESTfilename is the path/filename of our final output .xlsx file.
    copyAll(convertedList,itemList,exportFolder,DESTfilename)
    if clarg.unattended:
        sys.exit()

# DESTINATION FILE Existence Check or Create.
# NOTE: the reason we create this file is because we don't want any calculations in our final output,
# therefore, we are copying only values into a new file. No macro fears, no accidental corruption of any
# existing file.  Safer and cleaner this way, even though it takes extra time.
#
def makeOutput(itemList,exportFolder):
    try:
        currentxportFn = getCtrlVals()[3]
        if currentxportFn[-4:] != "xlsx":
            currentxportFn = currentxportFn + ".xlsx"
        if clarg.unattended:
            outFolderChk = u_outfolder
        else:
            outFolderChk = sysconfparse.get('folders', 'output_folder')
        LastMergeIni = outFolderChk + "\\LastXMerge.ini"
        confparse.read(LastMergeIni)
        confxportFn = confparse.get('export', 'exportfilename')
        if confxportFn != currentxportFn:
            confparse.set('export', 'exportfilename', currentxportFn)
            confxportFn = currentxportFn
            with open(LastMergeIni, 'w') as nnIni:
                confparse.write(nnIni)
        DESTfilename = exportFolder + "/" + confxportFn
        wb3 = xl.Workbook()
        wb3Sheet = wb3['Sheet']
        wb3Sheet.title = 'XMerge'
        ws3 = wb3.worksheets[0]
        text3.insert(tk.INSERT, '\n - Created ' + confxportFn + '.')
        window.update()
        wb3.save(filename = DESTfilename)
        wb3 = xl.load_workbook(DESTfilename, data_only=True) 
        ws3 = wb3.worksheets[0]
        incVar = 1
        for label in itemList:
            ws3.cell(row = 1, column = incVar).value = label
            incVar = incVar + 1
        wb3.save(DESTfilename)
        text4.insert(tk.INSERT, "\nCreated .xlsx file and header row.\n")
        window.update()
    except PermissionError:
        msg = "Close the output spreadsheet."
        endWithError(msg)

    return(DESTfilename)
#
# GET SOURCE FILE NAMES
#
def newOutFolder():
    files = []
    outxm = getCtrlVals()[1]        # Look for 'Change Output Folder' selection.
    ## Before doing anything else, check xmerge.ini to see if an output folder is 
    ## currently configured there.
    sysconfparse.read(XMergeIni)                                # Looking in xmerge.ini
    outFolderChk = sysconfparse.get('folders', 'output_folder') # get the existing value
    ofExist = os.path.isdir(outFolderChk)                       # find out if it exists
    if ofExist == False or outFolderChk == "": # Configured output folder isn't there or xmerge.ini doesn't have one.
        outFilesChkBox.select()     # Check the 'Change Output Folder' box.
        outxm = 1
    userHome = os.path.expanduser("~")
    if outxm == 1:                                  # Output folder selection called for. Open File dialog.
        text3.delete("1.0", tk.END)
        text3.insert("1.0", "Choose the folder where XMerge should put output file.\nYou might want to create a new sub-folder for these.")
        if outFolderChk != "":
            text3.delete("1.0", tk.END)
            text3.insert("1.0", "Verify this folder is where you want to put your merged data file.")
            window.update()
            exportFolder = filedialog.askdirectory(
                initialdir=(outFolderChk), title="This is the previous OUTPUT folder. Use it again, or navigate to another one.")
        else:
            text3.delete("1.0", tk.END)
            text3.insert("1.0", "Select or Create a folder for your output files.")
            window.update()
            exportFolder = os.path.normpath(filedialog.askdirectory(
                initialdir=(userHome), title="Select or Create a folder for your output files."))
        if exportFolder == "":
            keepGoing = messagebox.askyesno("No folder selected.", "Start over? Y/N")
            if keepGoing == True:
                main()
            else:
                sys.exit()
        outFolderChk = os.path.normpath(exportFolder)
        sysconfparse.set('folders','output_folder', outFolderChk)
        with open(XMergeIni, 'w') as csvIni:
            sysconfparse.write(csvIni)
        outxm = 0
        outFilesChkBox.deselect()
        chkIniexist = os.path.isfile(exportFolder + "\\LastXMerge.ini")     #  look for LastXMerge.ini.
        if chkIniexist == True:
            thisMerge = os.path.normpath(exportFolder + "\\LastXMerge.ini")
            runThis = messagebox.askyesno("LastXMerge.ini found.", "Use it as it is?")
            if runThis == False:
                newone = messagebox.askyesno("Please Choose","Y creates new .ini. N exits.")
                if newone == False:
                    setup()
                else:
                    newLastXMerge()
                    setup()
                    messagebox.showinfo("New LastXMerge.ini Created","Your new .ini file was created."
                    "\n\nDouble-check your settings.\n\nThen press 'Get New Data' or 'Go'.")
                window.mainloop()
            else:
                confparse.read(thisMerge)     # Read from LastXMerge.ini in the new folder.
                confxportFn = confparse.get('export', 'exportfilename')
                if confxportFn[-4:] != "xlsx":
                    confxportFn = confxportFn + ".xlsx"
                colNameVar.set(confparse.get("userselections","columnselection"))
                xportFn.set(confxportFn)     # puts updated Export Filename in textbox.
    return outFolderChk
    
    #
    # CREATE A NEW LastXMerge.ini FILE IN THE OUTPUT FOLDER.
    #
def newLastXMerge():
    if clarg.unattended:
        sys.exit()  ## You can't select files if you're not here.  So all we can do is exit.
    else:
        outFolderChk = sysconfparse.get('folders', 'output_folder')
    if outFolderChk != "":
        LastMergeIni = outFolderChk + "\\LastXMerge.ini"
        confparse["files"]={}
        confparse["userselections"]={"columnselection":"all","headerrow":"auto","appendfilename":"0"}
        confparse["folders"]={"data_folder1":""}
        xpfn = getCtrlVals()[3]
        confparse["export"]={"exportfilename":xpfn}
        with open(LastMergeIni, 'w') as LMIni:
            confparse.write(LMIni)
    else:
        outFilesChkBox.select()     # Check the 'Change Output Folder' box.
        outxm = 1
        main()
    #
    # Now get the source filenames.
    #
def browseFiles():
    files = []
    # Look in the system .ini file for the last Output Folder we used.
    if clarg.unattended:
        outFolderChk = u_outfolder
    else:
        outFolderChk = sysconfparse.get('folders', 'output_folder')
    LastMergeIni = outFolderChk + "\\LastXMerge.ini"
    outxm = getCtrlVals()[1]        # Get the controls status value for 'Change Output Folder' selection.
    chkIniexist = os.path.isfile(LastMergeIni)    #  look for LastXMerge.ini in the last place we worked.
    if chkIniexist == False:
        setup()                      # Make a new LastXMerge.ini if necessary.
    # Otherwise, use the one that was there.
    # Grab the list of paths and filenames from LastXMerge.ini.
    datchk = {k:v for k,v in confparse['files'].items()}
    dataFolder = confparse.get('folders','data_folder1')     # Get the source data folder name.
    exportFolder = outFolderChk                             # Get the export destination folder name.
    userHome = os.path.expanduser("~")                       # Get the current user home folder name.
    text3.delete("1.0", tk.END)
    text3.insert("1.0", "Select Source ATTOUT .txt, .xlsx or .csv files.")
    # When there are no files in LastXMerge.ini: 
    if datchk == {}:
        srcSelIter = 1
        # 'srcSelIter' is our incrementing variable to sequentially number the entries
        # in LastXMerge.ini so each filename gets a unique key in the config file.
        moreVar = 1
        while moreVar >= 1:
            srcSelection = filedialog.askopenfilenames(
                initialdir=dataFolder,
                title="Select Source files.",
                multiple=True,
                )
            if srcSelection != "":
                for fn in srcSelection:
                    confparse.set('files','file'+str(srcSelIter),os.path.normpath(fn))
                    srcSelIter = srcSelIter + 1
                dataFolder = os.path.dirname(srcSelection[0])
                confparse.set('folders','data_folder'+str(moreVar),os.path.normpath(dataFolder))
                with open(LastMergeIni, 'w') as csvIni:
                    confparse.write(csvIni)
            else:
                keepGoing = messagebox.askyesno("No file was selected.", "Continue? Y/N")
                if keepGoing == True:
                    window.mainloop()
                else:
                    sys.exit()
            moreFiles = messagebox.askyesno("Select More Files","Do you want more files? Y/N")
            if moreFiles == True:
                moreVar += 1
            else:
                moreVar = 0
            
    # When there ARE files in LastXMerge.ini.
    datchk = {k:v for k,v in confparse['files'].items()}
    datchkFiles = []
    for k in datchk:
        datchkFiles.append(datchk[k])                       # one for Windows
    firstFile = datchkFiles[0]
    if firstFile == "":
        keepGoing = messagebox.askyesno("No file was selected.", "Continue? Y/N")
        if keepGoing == True:
            pass
        else:
            sys.exit()
    confparse.set('export', 'exportfilename', getCtrlVals()[3])
    colNameChoice = getCtrlVals()[0]    # Get current Columns Selection checkbox value.
    hdrRowChoice = getCtrlVals()[4]     # Get Header Row Number textbox value.
    confparse.set("userselections","columnselection",colNameChoice)
    confparse.set("userselections","headerrow",hdrRowChoice)
    with open(LastMergeIni, 'w') as csvIni:
        confparse.write(csvIni)
    text3.delete("1.0", tk.END)
    window.update()
    fileNam = firstFile
    updateWinText()
    return firstFile,exportFolder,datchkFiles
#
# COPY SOURCE FILES TO EXPORT FOLDER. These are the raw files, prior to conversion.
#
def exCOPYcontinue(exportFolder,datchkFiles):
    tempFolder = os.path.join(exportFolder, "temp")
    # if the temp folder exists, skip creation
    if not os.path.isdir(tempFolder):
        os.makedirs(tempFolder)                 # Create a temporary folder for working files inside selected folder.
    sx = 0                                  # initialize a counter
    try:
        for f in datchkFiles:                   # 'datchkFiles' is filled by 'browseFiles()' and is full paths.
            if f != "":                         # It is empty if Cancel was pressed during 'browseFiles()'
                if sx == 0:                     # or if the source file was never selected.
                    text4.delete('1.0',tk.END)
                text4.insert('1.0', "Copying " + os.path.basename(f) + '. \n')
                window.update()
                time.sleep(.5)
                fbase = os.path.basename(f)
                if os.path.isfile(os.path.join(tempFolder,fbase)):
                    fsplit = os.path.splitext(fbase)
                    newf = fsplit[0] + str(sx)
                    fbase = newf + fsplit[1]
                fdest = os.path.join(tempFolder,fbase)
                shutil.copyfile(f,fdest)     # Copy the file to the temp folder inside the Output folder.
                sx = sx + 1                     # Increment the counter.
    except FileNotFoundError:
        keepGoing = messagebox.askyesno("File Not Found",os.path.basename(f) + " was not found. Exit?")
        if keepGoing == False:
            text4.delete('1.0',tk.END)
            text4.insert(tk.INSERT, "You need to check the contents of LastXMerge.ini"
            " in your Output Folder.  One or more of your source files' name has changed, or it is gone."
            "\n\nYou will need to exit XMerge. Then open LastXMerge.ini with Notepad and edit it."
            "\n\nOR you can use the 'Get New Data' button to reselect sources.")
            window.update()
            window.mainloop()
        else:
            sys.exit()
    text3.delete('1.0',tk.END)
    text3.insert(tk.END, "Finished copying " + str(sx) + ' files to the Temp Folder. ')
    window.update()
#
#  NOW THAT SOURCE FILES ARE IN THE OUTPUT FOLDER, IT'S TIME TO CONVERT:
#  Create lists of existing files of the various file-types we want to deal with.
#
def dataCompile():
    
    text3.delete('1.0',tk.END)
    text3.insert(tk.END, "Processing data files.  Standby.")
    window.update()
    text4.delete('1.0',tk.END)
    text4.insert(tk.INSERT, "Follow progress here.")
    

def initialize(exportFolder):
    tempFolder = os.path.join(exportFolder, "temp")
    xlsext = []                                                    # a list for files with .xls extensions. 
    xlsxext = []                                                   # a list for files with .xlsx extensions. 
    txtext = []                                                    # a list for files with .txt extensions. 
    csvext = []                                                    # a list for files with .csv extensions. 
    allEntr = []
    with os.scandir(tempFolder) as listOfEntries:
        for entry in listOfEntries:
            if entry.is_file():
                strEntry = entry.name
                allEntr.append(strEntry)
                splitFile = strEntry.rsplit('.',1)
                if splitFile[1] == 'xls':
                    xlsext.append(strEntry)
                if splitFile[1] == 'xlsx':
                    xlsxext.append(strEntry)
                if splitFile[1] == 'txt':
                    txtext.append(strEntry)
                if splitFile[1] == 'csv':
                    csvext.append(strEntry)
    return xlsext, xlsxext, txtext, csvext, allEntr

#
#  DEFINE .csv FILE HANDLER FUNCTION
# This is for comma-delimited files (hence the .csv extension).
#
def csvFileCnvrt(exportFolder,csvext,convertedList):
    tempFolder = os.path.join(exportFolder, "temp")
    text4.insert('1.0', "Converting .csv files to .xlsx.\n")
    for csvFile in csvext:
        csvFile = tempFolder + "/" + csvFile
        csvwb = xl.Workbook()
        csvws = csvwb.active
        with open(csvFile) as f:
            f.seek(0)
            if "mount" in csvFile:
                guts = csv.reader(f,delimiter = ';')
            else:
                guts = csv.reader(f,delimiter = ',')
            for row in guts:
                csvws.append(row)
        saveFile = csvFile.replace('csv','xlsx')
        csvwb.save(saveFile)
        convertedList.append(saveFile)
        text4.insert('1.0', "Converted ==> " + os.path.basename(saveFile)+"\n")
        window.update()
    return convertedList
#
#  END .csv FILE HANDLER
#
#  DEFINE .txt FILE HANDLER FUNCTION
#  These files are Tab-delimited. Typically from AutoCAD's ATTOUT routine.
#
def txtFileCnvrt(exportFolder,txtext,convertedList):
    tempFolder = os.path.join(exportFolder, "temp")
    text4.insert('1.0', "Converting .txt files to .xlsx.\n")
    for t in txtext:
        tf = t
        t = tempFolder + "/" + t
        csvwb = xl.Workbook()
        csvws = csvwb.active
        with open(t) as f:
            guts = csv.reader(f, delimiter='\t')
            for row in guts:
                csvws.append(row)
        saveFile = t.replace('txt','xlsx')
        csvwb.save(saveFile)
        convertedList.append(saveFile)
        text4.insert('1.0', "Converted ==> " + os.path.basename(saveFile)+"\n")
        window.update()
    return convertedList
#
#  END .txt FILE HANDLER
#
#  DEFINE .xls FILE HANDLER FUNCTION
#
def xlsFileCnvrt(exportFolder,xlsext,convertedList):
    import pyexcel
    import pyexcel_xls
    import pyexcel_xlsx
    tempFolder = os.path.join(exportFolder, "temp")
    text4.insert('1.0', "Converting .xls files to .xlsx.\n")
    for xlsFile in xlsext:
        xf = xlsFile
        xlsFile = tempFolder + "/" + xlsFile
        xlsxFile = xlsFile.replace('xls','xlsx')             # .xls file.  We need an .xlsx file.
        pyexcel.save_book_as(file_name=xlsFile,
                    dest_file_name=xlsxFile)
        convertedList.append(xlsxFile)
        text4.insert('1.0', "Converted ==> " + os.path.basename(xlsxFile)+"\n")
        window.update()
    return convertedList
#
#  END .xls FILE HANDLER
#
# Column Name Row Number Determiner #
# Many of the spreadsheets in our work have a single cell on Row 1 with a title in it.
# Row 2 then contains the header row with column names in it.
# If the user has not specified a specific number of rows to skip, this function
# will check for the existence of a Title above the Header Row.
#
def findHeaderRow(filename,sheetname):
    hdrrwFnVar = getCtrlVals()[4]               # lookup the Header Row value.
    if str.lower(hdrrwFnVar) != 'auto':
        if len(hdrrwFnVar) == 1:
            hrow = int(hdrrwFnVar)
        else:
            msg = "Must be 'auto' or a number\nbetween 1 & 9."
            endWithError(msg)

    if str.lower(hdrrwFnVar) == 'auto':
        wb1 = xl.load_workbook(filename, data_only=True)
        if sheetname == None:
            sheetname = 0                       # Check the first default sheet when 'sheetname' is not specififed.
        ws1 = wb1.worksheets[sheetname]
        ii = range(1,5)                         # We will check 4 cells to determine whether it is a title or header row.
        cellAddr1 = []                          # Set up to check row 1.
        for i in ii:
            thisOne = ws1.cell(row = 1, column = i).value
            if thisOne != " " and thisOne != None:
                cellAddr1.append(thisOne)
        if len(cellAddr1) < 4:                  # Row 1 had fewer than 4 cells with data in them.
            ii = range(1,5)                     # Check 4 cells in row 2.
            cellAddr2 = []                      # Set up to check row 2.
            for i in ii:
                thisOne = ws1.cell(row = 2, column = i).value
                if thisOne != " " and thisOne != None:
                    cellAddr2.append(thisOne)
            if len(cellAddr2) < 4:              # If row 2 has less than 4 full cells, default back to row 1 as header.
                hrow = 1                        # Maybe this sheet simply has fewer than 4 columns.  (??)
            else:
                hrow = 2                        # Otherwise, we'll use row 2 as the header row because row 1 is a title.
        else:
            hrow = 1                            # Because row 1 had at least 4 full cells, it is likely the header row.
    return hrow

#
#  DEFINE COLUMN NAMES GRABBER FUNCTION
#  We are using the first source file's column names as the definition for the Ouput file.
#
def makeColsList(filename):
    with warnings.catch_warnings(record=True):         # these two lines are a sad way to stop getting a 
        warnings.simplefilter("always")                # warning about the lack of a default style
        wb1 = xl.load_workbook(filename, data_only=True) 
        ws1 = wb1.worksheets[0] 
        mr = ws1.max_row                               # Get the total number of rows in the Source spreadsheet.
        HeaderRow = findHeaderRow(filename,0)
        columnsDB = []
        SrcColIndex = {}                                # SrcColIndex will be the dictionary of Source column names.
        Current  = 0                                   # Get all of the column names from the Source file.
        for itrCol in ws1.iter_cols(min_row=HeaderRow, max_col=ws1.max_column):   # Store them in the dictionary. Key is the Name, value is the column number.
            if itrCol[0].value != None:
                SrcColIndex[itrCol[0].value] = Current
            Current += 1
        for k in SrcColIndex:
            columnsDB.append(k)
    return columnsDB
#
# CONFIGURATION DATA GRABBER (Use this when there's a Column Name Configuration spreadsheet in use.)
#
def colConfig():                                        # open the Configuration workbook ColumnNames.xlsx .
    configFile = manageColNames()
    text1.insert(tk.INSERT,"\nColumn Config File: " + str(configFile))
    window.update()
    confwb = xl.load_workbook(configFile, data_only=True)  # Load the Configuation workbook.
    try:
        confws = confwb["MERGE"]                              # Use the sheet labeled with the MERGE label.
    except KeyError as kerr:                               # DataColumnsNames may not be up to date.
        messagebox.showinfo("Column Name Issue.",kerr)
        text4.delete('1.0',tk.END)
        text4.insert(tk.INSERT, "ColumnNames.xlsx needs attention.")
        editNow = messagebox.askyesno("Edit ColumnNames Now?", "Yes to edit now, No to exit.")
        if editNow == True:
            editColNams()
        else:
            sys.exit()
    columnsDB = []                                          # Set up the column name list.
    tcnMaxR = confws.max_row                               # Find out how many rows there are. (Not just Col A)
    for cn in range(1,tcnMaxR+1):                          # Iterate through those rows and 
        colNameAddr = confws.cell(row = cn, column = 1)    # memorize the cell address, then
        text_string = colNameAddr.value                    # set about making sure the name is normalized unicode text.
        if text_string != None:                            # This ignores rows with empty column A but text in B or . . . 
            cleanedColName = unicodedata.normalize("NFKD",text_string)
            columnsDB.append(cleanedColName)                    # append the column names to the list.
    return columnsDB
#
# COPY DATA FROM SOURCE FILES TO DESTINATION FILE
#       convertedList brings in the filenames for source files.
#
def copyAll(convertedList,itemList,exportFolder,DESTfilename):
    tempFolder = os.path.join(exportFolder, "temp")
    ColNames = itemList                                         # Rename itemList to colNames for clarity.
    sumRow = 0
    SrcMissCol = ""                                   # Set variable to show Source had Missing Columns.
    for sn in convertedList:
        text4.delete('1.0',tk.END)
        text4.insert(tk.INSERT, "\nProcessing Sources: \n" + os.path.basename(sn) + ".")
        src1 = xl.load_workbook(sn, data_only=True)             # Get source file as 'src1'.
        srcSh1 = src1.worksheets[0]                             # Set source sheet to first sheet.
        # COUNT ONLY ROWS THAT ACTUALLY HAVE DATA IN THEM.
        # openpyxl's max_row() function returns the index of the last row, not the last row with data in it.
        # That means if there are tons of blank rows, they all get copied to the Output data file, which
        # is not optimal.  So here we create an index of rows that are NOT empty in a list named srcRowIdx.
        srcSh1_mr = 0                                           
        srcRowIdx = []
        for row in srcSh1:
            if not all([(cell.value is None or cell.value == "") for cell in row]):
                srcSh1_mr += 1
                copyRow = row[1].row
                srcRowIdx.append(copyRow)
        text4.insert(tk.INSERT, "\nData is in " + str(srcSh1_mr - 1) + " rows.")
        window.update()
        out1 = xl.load_workbook(DESTfilename, data_only=True)   # Get destination file as 'out1'.
        outSh1 = out1.worksheets[0]                             # Set destination sheet to first sheet.
        outSh1_mr = outSh1.max_row                              # Get the max row for the destination sheet
        outRow1 = outSh1_mr + 1                                 # Set the initial destination row to write data into.
        jWrite = 1                                              # Set an initial column number to write data into.
        #
        # We want to record the index of the source file's column names because this allows us to import data from files
        # that have the same column names, but in a different order from the destination file. As long as the column name
        # matches exactly, we'll get the data from it no matter where it is in the sheet.
        #
        # We need to get the index number of the SOURCE column names. ColNames (itemList) has what we WANT. This index is what we HAVE.
        SrcColIndex = {}                                  # SrcColIndex is the dictionary of Source column names. Key is the Name, value is the column number.
        SrcHeader = findHeaderRow(sn,0)                   # Header is the row where column names are expected to be found.
        Current  = 0                                      # Get all of the column names from the Source file.
        for itrCol in srcSh1.iter_cols(min_row=SrcHeader, max_row = SrcHeader, max_col=srcSh1.max_column):   # Store them in the dictionary.
            SrcColIndex[itrCol[0].value] = Current
            Current += 1
        SrcColIndex['SRCFILE'] = Current                  # Add an index entry for the SRCFILE name at the end of the dictionary.
        DestColIndex = {}
        Current  = 0                                      # Get any column names from the Destination file.
        for itrCol in outSh1.iter_cols(min_row=1, max_row = 1, max_col=outSh1.max_column):   # Store them in the dictionary.
            DestColIndex[itrCol[0].value] = Current
            Current += 1
        try:
            for n in ColNames:                                              # n is a Column Name .
                text4.insert(tk.INSERT, "\n - Filling " + n + ".")          # Provide user feedback.
                window.update()
                try:
                    jRead = SrcColIndex[n] + 1                                  # jRead is the column index from the Source's same-named column 
                                                                                # (key: column name, value: column number).
                except KeyError:                                                # If Source doesn't have this column, flag it with 99999
                    jRead = 99999                                               # That allows us to provide it with an empty string.
                jWrite = DestColIndex[n] + 1                                    # Memorize the column index from the Destination column. Assume nothing.
                outRow = outRow1                                                # Not sure why I did this, I like to think there was a reason.
                srcRowsLen = len(srcRowIdx)
                for rowVar in range(SrcHeader, srcRowsLen):                     # From the Header row + 1 to the SOURCE sheet MAX row number + 1.
                            # instead of max_row, we use a list of row index numbers created above.
                    rowNum = srcRowIdx[rowVar]
                    if jRead == 99999:
                        SrcMissCol = "y"
                        cValue = ""                                             # If Source didn't have the column, provide a empty string value.
                        outSh1.cell(row = outRow, column = jWrite).value = cValue # writing the null value to destination excel file
                    else:
                        if n == 'SRCFILE':
                            cValue = pathlib.PurePath(sn).stem                  # Grab just the filename w/o the extension.
                        else:
                            cValue = srcSh1.cell(row = rowNum, column = jRead).value               # read cell value from source excel file
                        if cValue == "<>":
                            cValue = ""
                        outSh1.cell(row = outRow, column = jWrite).value = cValue # writing the read value to destination excel file
                        # if we are at the 
                    outRow = outRow + 1                                         # increment the destination row number.
        except Exception as e:
            endWithError(str(e) + "\nCheck your Header Row number.")
        out1.save(DESTfilename)
    ##
    do_csv = getCtrlVals()[6]
    if do_csv == 1:
        ixf = DESTfilename
        ocf = DESTfilename.replace('xlsx','csv')
        xlsx_to_csv(ixf,ocf)
        csvVar.set(0)
    ## The simplest way to deal with NOT producing an .xlsx is to just delete it afterwards.
    ## Everything happens in the .xlsx format, so we will build it, then if we don't want it, delete.
    do_xlsx = getCtrlVals()[7]
    if do_xlsx == 0:
        os.remove(DESTfilename)
        if do_csv == 1:
            DESTfilename = ocf
        xlsxVar.set(1) # We reset this to ON because .xlsx is the default output format.
    if do_csv == 0 and do_xlsx == 0:
        DESTfilename = "No Output File"
    sumRow = sumRow + outRow
    tempFolder = os.path.join(exportFolder, "temp")
    shutil.rmtree(tempFolder)
    text4.delete('1.0',tk.END)
    text4.insert(tk.INSERT, os.path.basename(DESTfilename) + " was created in \n" + exportFolder + ".\nThere are " + str(sumRow-2) + " records in the new spreadsheet."
    "\n\nThe temp folder was deleted.")
    if do_csv == 1:
        text4.insert(tk.INSERT, "\n\nThe .csv file was created.")
    if SrcMissCol == "y":
        text4.insert(tk.INSERT, "\n\nAt least one Source file had a missing column.\n"
        "For that source, cells in that column were filled with empty strings.")
    text3.delete('1.0',tk.END)
    text3.insert(tk.INSERT,os.path.basename(DESTfilename) + " was created.")
#
# PURGE ALL FILES FROM THE OUTPUT FOLDER
#
def purgeExp():
    ofcchk = getCtrlVals()[2]
    if ofcchk == 1:
        text4.delete('1.0',tk.END)
        text4.insert(tk.INSERT, "You are about to delete every file in your most recent Output folder.")
        window.update()
        keepGoing = messagebox.askyesno("Purge Output", "Are you sure? Y/N")
        if keepGoing == True:
            exportFolder = sysconfparse.get('folders', 'output_folder')
            tmpFolder = os.path.join(exportFolder, "temp")
            if os.path.isdir(tmpFolder) == True:
                shutil.rmtree(tmpFolder)
            with os.scandir(exportFolder) as listOfEntries:
                for entry in listOfEntries:
                    os.remove(entry)
            newLastXMerge()
            text4.delete('1.0',tk.END)
            text4.insert(tk.INSERT, "Your Output folder is now empty. A new LastXMerge.ini was created.")
            window.update()
        else:
            # sys.exit()
            text4.delete('1.0',tk.END)
            text4.insert(tk.INSERT, "No files were deleted.")
            text3.delete('1.0',tk.END)
            text3.insert(tk.END, "No files were deleted.")
            window.update()
    OFClearChkBox.deselect()
#
# THE DEFINITIVE EXIT FUNCTION.
#
def exit():
    ofcchk = getCtrlVals()[2]   # Is 'Clear Output Folder' checked?
    if ofcchk == 1:
        purgeExp()
    sys.exit()
#
# END WITH ERROR MESSAGE
#
def endWithError(msg):
    messagebox.showinfo("UhOh",msg)
    sys.exit()
#
#  Somebody pushed the 'Get New Data' button. GET NEW DATA SOURCE FILES.
#
def newData():  # Resets everything as if there is no data file, then calls main()
    sysconfparse.read(XMergeIni)            # Use the system-wide ini file.
    outFolderChk = sysconfparse.get('folders', 'output_folder')    # from XMerge.ini
    LastMergeIni = outFolderChk + "\\LastXMerge.ini"
    # If Change Output Folder is checked, leave this LastXMerge.ini file alone.
    # If it's not checked, we want new data in the same Output Folder.
    # So, then, we need to clear the filenames from LastXMerge.ini in this 
    # Output folder.
    outxm = getCtrlVals()[1]        # Check the controls status value for Change Output folder selection.
    ofcchk = getCtrlVals()[2]       # Look at 'Clear Output Folder' checkbox.
    if ofcchk == 1:                 # It's checked, we know what to do.
        text4.delete('1.0',tk.END)
        text4.insert(tk.INSERT, "Purging the Exports Folder.")
        purgeExp()                  # Go to the Clear Output function.
    text1.delete("1.0", tk.END)
    text4.delete("1.0", tk.END)
    text4.insert(tk.INSERT, genInfo)
    text3.delete("1.0", tk.END)
    text2.delete("1.0", tk.END)
    window.update()
    if outxm == 1:
        text3.delete('1.0',tk.END)
        text3.insert(tk.END, "Clearing and getting new job data.")
        outFolderChk = newOutFolder()
    if outxm == 0:
        chkIniexist = os.path.isfile(LastMergeIni)      #  look for LastXMerge.ini.
        if chkIniexist == False:                        # There wasn't one.
            confparse["files"]={}                       # Set up the basic settings.
            confparse["userselections"]={"columnselection":""}   #
            confparse["folders"]={"data_folder1":""}     #
            xpfn = getCtrlVals()[3]
            confparse["export"]={"exportfilename":xpfn} # Use whatever filename is on the screen.
            with open(LastMergeIni, 'w') as LMIni:      #  Create a new LastXMerge.ini file.
                confparse.write(LMIni)                  #
        # Now there is a LastXMerge.ini, either old or new.
        confparse.read(LastMergeIni)            # Read it.
        expFileName = getCtrlVals()[3]
        if expFileName[-4:] != "xlsx":          # if not an .xlsx extension, supply it.
            expFileName = expFileName + ".xlsx"
        xportFn.set(expFileName)
        confparse.set("export","exportfilename",expFileName)
        datchk = {k:v for k,v in confparse['files'].items()}
        for k in datchk:
            confparse.remove_option('files',k)      # Purge any and all filename entries.
        confparse.set('folders','data_folder1',"")   # Reset the Output Folder to blank.
        with open(LastMergeIni, 'w') as csvIni: # Save the edits.
            confparse.write(csvIni)
    main()      # Re-run the main() function, which sends us back to gathering sources.
                # That gathering happens in browseFiles().

def xlsx_to_csv(ixf,ocf):
    inputxlsx = xl.load_workbook(ixf)
    ixsheet = inputxlsx.active
    with open(ocf, 'w', newline="") as thecsv:
        csv_writer = csv.writer(thecsv)
        for r in ixsheet.iter_rows():
            csv_writer.writerow([cell.value for cell in r])

#
#  EVERYTHING SOUTH OF HERE defines the tKinter windows stuff  #
#
#
window = tk.Tk()  # Create the root window.  'root' is the common name, but I named this one 'window'.
window.title("XMerge: Data Source File Merger")  # Set window title
winWd = 1000  # Set window size and placement
winHt = 800
x_Left = int(window.winfo_screenwidth() / 2 - winWd / 2)
y_Top = int(window.winfo_screenheight() / 2 - winHt / 2)
window.geometry(str(winWd) + "x" + str(winHt) + "+{}+{}".format(x_Left, y_Top))
window.config(background="white")  # Set window background color
window.columnconfigure(0, weight=0)
window.columnconfigure(1, weight=0)
window.columnconfigure(2, weight=1)
window.columnconfigure(3, weight=0)
window.columnconfigure(4, weight=0)
window.columnconfigure(5, weight=0)
window.columnconfigure(6, weight=1)
window.rowconfigure(0, weight=1)
window.rowconfigure(1, weight=1)
window.rowconfigure(2, weight=1)
window.rowconfigure(3, weight=1)
window.rowconfigure(4, weight=1)
window.rowconfigure(5, weight=1)
window.rowconfigure(6, weight=1)
label_file_explorer = tk.Label(
    window,  # Create a File Explorer label
    text="XMerge: Data Source File Merger",
    width=winWd,
    font=18,
    justify="center",
    fg="navy",
    bg="light blue"
)

#
# FEATURE NOT READY
#
def featureNotReady():
    messagebox.showinfo(title='Not Yet', message='That feature is not ready.')
#
# DEFINE THE ABOUT WINDOW
#
def aboutWindow():
    aw = tk.Toplevel(window)
    aw.title("About XMerge")
    awinWd = 400  # Set window size and placement
    awinHt = 400
    x_Left = int(window.winfo_screenwidth() / 2 - awinWd / 2)
    y_Top = int(window.winfo_screenheight() / 2 - awinHt / 2)
    aw.config(background="white")  # Set window background color
    aw.geometry(str(awinWd) + "x" + str(awinHt) + "+{}+{}".format(x_Left, y_Top))
    aw.iconbitmap(path_to_dat + './ico/XMergeicon.ico')
    awlabel = tk.Label(aw, font=18, text ="About XMerge " + version)
    awlabel.grid(column=0, columnspan=3, row=0, sticky="n")  # Place label in grid
    aw.columnconfigure(0, weight=1)
    aw.rowconfigure(0, weight=1)
    aboutText = tk.Text(aw, height=20, width=170, bd=3, padx=10, pady=10, wrap=tk.WORD, font=nnFont)
    aboutText.grid(column=0, row=1)
    aboutText.insert(tk.INSERT, "This tool converts and merges multiple flat source files (.csv, .txt, .xls, .xlsx) into one .xlsx file." 
"\n\nCheck out Help for more details.\n\nYour XMerge installation and supporting files are located at:\n\n" + path_to_dat + "\n"
"\nStart with the 'Get New Data' button to the left.\n\n- Greg Sanders, aka Dr.Gerg\n"
"\nXMerge is written in Python and compiled using PyInstaller.\nInno Setup Compiler builds the Windows installer.\n\n"
"https://www.drgerg.com\nhttps://github.com/drgerg/XMerge")
#
# DEFINE THE HELP WINDOW
#
def helpWindow():
    hw = tk.Toplevel(window)
    hw.tk.call('tk', 'scaling', 1.0)    # This prevents the text being huge on hiDPI displays.
    hw.title("XMerge Help")
    hwinWd = 600  # Set window size and placement
    hwinHt = 600
    x_Left = int(window.winfo_screenwidth() / 2 - hwinWd / 2)
    y_Top = int(window.winfo_screenheight() / 2 - hwinHt / 2)
    hw.config(background="white")  # Set window background color
    hw.geometry(str(hwinWd) + "x" + str(hwinHt) + "+{}+{}".format(x_Left, y_Top))
    hw.iconbitmap(path_to_dat + './ico/XMergeicon.ico')
    hwlabel = HTMLLabel(hw, height=3, html='<h2 style="text-align: center">XMerge Help</h2>')
    hw.columnconfigure(0, weight=1)
    helpText = HTMLScrolledText(hw, height=44, padx=10, pady=10, html=RenderHTML(path_to_dat + "\XMerge_Help.html"))
    hwlabel.grid(column=0, row=0, sticky="NSEW")  # Place label in grid
    helpText.grid(column=0, row=1, ipadx=10, ipady=10, sticky="NSEW")
#
def manageColNames():
    outFolderChk = sysconfparse.get('folders', 'output_folder')
    colPath = str(outFolderChk + '/ColumnNames.xlsx')
    coln = pathlib.PurePath(colPath)
    fileChk = os.path.isfile(coln)  # Look for the file.
    if fileChk == False:            # copy it if absent.
        shutil.copy(path_to_dat + '\ColumnNames.xlsx',outFolderChk)
        text4.insert(tk.INSERT, "\nColumnNames.xlsx was missing. Added it to Output Folder.")
        # Ask if we need to edit the freshly copied file. Almost certainly yes.
        doEdit = messagebox.askyesno("ColumnNames.xlsx was just copied over.","Do you probably need to edit it.")
        if doEdit == True:
            editColNams()
    return coln

def editColNams():
    coln = manageColNames() # The first part of manageColNames() provides our path and filename.
    colNameVar.set("Conf")
    os.startfile(coln)      # Open the spreadsheet using the system default application.
    messagebox.showinfo("Editing ColumnNames","After editing, hit 'GO' to do your Merge.")
    setup()
    window.mainloop()
#
# MENU AND MENU ITEMS
#
tk.Frame(window)
menu = tk.Menu(window)
window.config(menu=menu)
nnFont = Font(family="Segoe UI", size=10)          # Set the base font
fileMenu = tk.Menu(menu, tearoff=False)
fileMenu.add_command(label="Edit ColumnNames.xlxs", command=editColNams)
fileMenu.add_command(label="Exit", command=exit)
menu.add_cascade(label="File", menu=fileMenu)

editMenu = tk.Menu(menu, tearoff=False)
editMenu.add_command(label="Help", command=helpWindow)
editMenu.add_command(label="About", command=aboutWindow)
menu.add_cascade(label="Help", menu=editMenu)
#
window.iconbitmap(default = path_to_dat + '/ico/XMergeicon.ico')

genInfo = ("This tool converts and merges multiple source files into one .xlsx file."
"\n\nClick either the 'Get New Data' or the 'GO' button.\n\n"
"XMerge will remember the last job, and start there next time.\n\n"
"The Output Folder shown above is also the working folder. The source files you select "
"for merging will first be copied to a temp folder in the Output Folder.  When"
"XMerge finishes, that temp folder is removed, leaving a fresh sparkly shine."
"\n\nBy default, all columns in the first file processed will define the "
"header row for the merged file.  However, you can configure which columns are merged "
"by editing ColumnNames.xlsx.\n\n'Edit ColumnNames.xlsx' is under the File menu.")

#
# SET UP RADIO BUTTONS FOR COLUMN NAME SELECTION
#
# "controlsFrame" frames them nicely.
#
def getCtrlVals():
    colName = colNameVar.get() # 0 - All columns or Configured
    outxm = outxmVar.get()  # 1 - Change Output Folder
    ofcchk = OFClrVar.get() # 2 - Clear Output Folder
    xportFnVar = xportFn.get() # 3 - Export file name
    hdrrwFnVar = hdrrwFn.get() # 4 - Header Row Number 
    appendfilename = aFnVar.get() # 5 - Append Src Filenames
    csvCopyvar = csvVar.get() # 6 - Create a .csv output file
    xlsxCopyvar = xlsxVar.get() # 7 - Create a .xlsx output file
    return colName,outxm,ofcchk,xportFnVar,hdrrwFnVar,appendfilename,csvCopyvar,xlsxCopyvar

controlsFrame = tk.LabelFrame(window, text="Controls")             # larger frame to hold Radio Button frame
controlsFrame.grid(column=0, row=2, padx=10, sticky='nw')
colNameVar = tk.StringVar(value="all")
rbframe = tk.LabelFrame(controlsFrame, text="Columns Selection")  # Frame within a frame for Radio Buttons
rbframe.grid(column=0, row=2, padx=10, pady=10, sticky='n')

cs1 = tk.Radiobutton(rbframe, text = "All Columns", variable = colNameVar, value="all", command=getCtrlVals)  # define it
cs1.grid(column=0, row=1, sticky='nw')                                                                # place it
cs2 = tk.Radiobutton(rbframe, text ="Configured Cols", variable = colNameVar, value="conf", command=getCtrlVals) # repeat
cs2.grid(column=0, row=2, sticky='nw')
#
# Set up push-buttons
#
button_go = ttk.Button(controlsFrame, text="Get New Data", command=newData)         # "Get New Data" button
button_go.grid(column=0, row=8, padx=10, pady=10, sticky='n')                       # Place New Data button in grid
button_go = ttk.Button(controlsFrame, text="Go", command=main)                      # "Go" button
button_go.grid(column=0, row=13, padx=10, pady=10, sticky='n')                      # Place Go button in grid
button_exit = ttk.Button(controlsFrame, text="Exit", command=exit)                  # "Exit" button
button_exit.grid(column=0, row=14, padx=10, pady=10, sticky='n')                    # Place Exit button in grid
#
# Set up check boxes
#
outxmVar = tk.IntVar(value=1)
outFilesChkBox = tk.Checkbutton(controlsFrame,text='Change Output Folder', variable=outxmVar, onvalue=1, offvalue=0, command=getCtrlVals)      # define it
outFilesChkBox.grid(column=0, row=3, sticky='nw')                                                   # place it
OFClrVar = tk.IntVar(value=0)
OFClearChkBox = tk.Checkbutton(controlsFrame,text='Clear Output Folder', variable=OFClrVar, onvalue=1, offvalue=0, command=getCtrlVals)      # define it
OFClearChkBox.grid(column=0, row=4, sticky='nw')                                                   # place it
aFnVar = tk.IntVar(value=0)
aFnChkBox = tk.Checkbutton(controlsFrame,text='Append Src Filenames', variable=aFnVar, onvalue=1, offvalue=0, command=getCtrlVals)      # define it
aFnChkBox.grid(column=0, row=5, sticky='nw')                                                   # place it
csvVar = tk.IntVar(value=0)
csvChkBox = tk.Checkbutton(controlsFrame,text='Export to .csv File', variable=csvVar, onvalue=1, offvalue=0, command=getCtrlVals)      # define it
csvChkBox.grid(column=0, row=6, sticky='nw')                                                   # place it
xlsxVar = tk.IntVar(value=0)
xlsxChkBox = tk.Checkbutton(controlsFrame,text='Export to .xlsx File', variable=xlsxVar, onvalue=1, offvalue=0, command=getCtrlVals)      # define it
xlsxChkBox.grid(column=0, row=7, sticky='nw')                                                   # place it
xlsxChkBox.select()
#
# Set up confxportFn text entry box
#
xportFn = tk.StringVar(value = "XMerge_Export.xlsx")
xportFnLabel = tk.Label(controlsFrame, text="Export Filename:")
xportFnEntry = tk.Entry(controlsFrame, justify='center', textvariable = xportFn, width=20)
xportFnLabel.grid(column=0, row=9)
xportFnEntry.grid(column=0, row=10)
#
# Set up HeaderRowVar text entry box
#
hdrrwFn = tk.StringVar(value = "auto")
hdrrwFnLabel = tk.Label(controlsFrame, text="Header Row Number:")
hdrrwFnEntry = tk.Entry(controlsFrame, justify='center', textvariable = hdrrwFn, width=6)
hdrrwFnLabel.grid(column=0, row=11, pady=(10,0), sticky='s')
hdrrwFnEntry.grid(column=0, row=12, sticky='n')
# Set up text windows
#
text1 = tk.Text(window, height=6, width=150, wrap=tk.WORD, font=nnFont)
text2 = tk.Text(window, height=2, width=150, font=nnFont)
text3 = tk.Text(window, height=3, width=150, bg='light blue', font=nnFont)

text4Frame = tk.LabelFrame(window, text='Things You Should Know')
text4Frame.grid(column=4,row=2, padx=6, sticky='w')
# text4 = tk.Text(text4Frame, width=170, bd = 3, padx = 10, wrap=tk.WORD, font=nnFont)
text4 = tk.Text(text4Frame, height=28, width=80, padx = 10, wrap=tk.WORD, font=nnFont)
text4.grid(column=0, row=0)

label_file_explorer.grid(column=0, columnspan=7, row=0, sticky="n")  # Place label in grid

text1.grid(column=0, columnspan=7, row=1)
text2.grid(column=0, columnspan=7, row=6)
text3.grid(column=0, columnspan=7, row=7)


listFrame = tk.LabelFrame(window, text="Available")
listFrame.grid(column=1, row=2, padx=6, sticky='w')


dupeFrame = tk.LabelFrame(window, text="Dupes")
dupeFrame.grid(column=2, row=2, padx=6, sticky='w')

scrollbar = ttk.Scrollbar(window, orient='vertical')
scrollbar.grid(row=2, column=5, sticky='ns')
text4.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=text4.yview)

#
# EXPLANATION TEXT ON MAIN WINDOW.
#
# final = []
# dupesList = []
# gapsList = ""
# t4Label = ""
# showLists(final,dupesList,gapsList,t4Label)
text4.insert(tk.INSERT, genInfo)
# Instead of directly specifying a main() function, we let the window.mainloop() wait for a button press
# from one of the buttons we defined.  The function associated with the button defines what happens next.
#
setup()  # Run the setup function once per launch to make sure the basics are covered.
window.mainloop()  # Run the (not defined with 'def') main window loop.
